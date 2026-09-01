"""
Live Excel Updater Module.
Creates and updates an interactive Excel workbook with:
- Sheet 1: 'Live_Prices_&_Pivots' (Live prices, previous-day OHLCV, standard daily pivots)
- Sheet 2: 'Signals' (All detected reversal & pattern signals in rows)
Maintains in-memory openpyxl workbook for guaranteed persistence and uses COM automation
for real-time live window streaming when available.
"""

from __future__ import annotations

from datetime import datetime
import logging
import os
from pathlib import Path
import threading
import time
from typing import Any, Dict, List, Optional, TYPE_CHECKING
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from indicators.pivots import DailyPivots

if TYPE_CHECKING:
    from scanner.signal_engine import SignalEvent

logger = logging.getLogger(__name__)


class LiveExcelManager:
    """Manages real-time Excel workbook initialization, live cell updates, and COM synchronization."""

    SHEET1_NAME = "Live_Prices_&_Pivots"
    SHEET2_NAME = "Signals"

    def __init__(self, file_path: Path = config.EXCEL_FILE_PATH, auto_open: bool = config.EXCEL_AUTO_OPEN):
        self.file_path = Path(file_path)
        self.auto_open = auto_open
        self.pivots: Dict[str, DailyPivots] = {}
        self.symbol_row_map: Dict[str, int] = {}  # symbol -> 1-based row in Excel
        self._lock = threading.Lock()

        # In-Memory openpyxl Workbook & Sheets (immune to COM / deactivated license issues)
        self.wb: Optional[openpyxl.Workbook] = None
        self.ws1: Optional[openpyxl.worksheet.worksheet.Worksheet] = None
        self.ws2: Optional[openpyxl.worksheet.worksheet.Worksheet] = None

        # Cached live price updates for COM thread
        self._live_cache: Dict[str, Dict[str, Any]] = {}
        self._signals_cache: List[SignalEvent] = []

        # COM Application handle
        self._excel_app = None
        self._workbook_com = None
        self._sheet1_com = None
        self._sheet2_com = None
        self._com_initialized = False
        self._stop_com_thread = False
        self._bg_thread: Optional[threading.Thread] = None

    def initialize_workbook(self, pivots_map: Dict[str, DailyPivots]):
        """
        Builds the initial Excel template with formatting and populates initial pivot data.
        """
        self.pivots = pivots_map
        wb = openpyxl.Workbook()
        self.wb = wb

        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill_blue = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_fill_green = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        border_thin = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # -------------------------------------------------------------
        # 1. Sheet 1: Live Prices & Pivots
        # -------------------------------------------------------------
        ws1 = wb.active
        ws1.title = self.SHEET1_NAME
        ws1.views.sheetView[0].showGridLines = True
        self.ws1 = ws1

        headers1 = [
            "Symbol", "LTP", "Change %", "Volume",
            "PDO", "PDH", "PDL", "PDC", "PDV",
            "Pivot (PP)", "R1", "R2", "R3", "S1", "S2", "S3",
            "Last Updated"
        ]
        ws1.append(headers1)

        # Style Header Row
        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill_blue
            cell.alignment = align_center

        row_idx = 2
        for symbol in sorted(pivots_map.keys()):
            p = pivots_map[symbol]
            self.symbol_row_map[symbol] = row_idx

            ws1.append([
                symbol,
                p.pdc,       # Initial LTP defaults to PDC
                0.00,        # Initial Change %
                p.pdv,       # Volume
                p.pdo, p.pdh, p.pdl, p.pdc, p.pdv,
                p.pp, p.r1, p.r2, p.r3, p.s1, p.s2, p.s3,
                datetime.now().strftime("%H:%M:%S")
            ])

            # Apply light borders & number formats
            ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws1.cell(row=row_idx, column=1).font = Font(name="Calibri", size=10, bold=True)
            for c in range(2, len(headers1) + 1):
                cell = ws1.cell(row=row_idx, column=c)
                cell.border = border_thin
                cell.alignment = align_right
                if c in [2, 5, 6, 7, 8, 10, 11, 12, 13, 14, 15, 16]:
                    cell.number_format = "#,##0.00"
                elif c == 3:
                    cell.number_format = "+0.00%;-0.00%;0.00%"
                elif c in [4, 9]:
                    cell.number_format = "#,##0"

            row_idx += 1

        # Auto-fit columns
        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # -------------------------------------------------------------
        # 2. Sheet 2: Signals
        # -------------------------------------------------------------
        ws2 = wb.create_sheet(title=self.SHEET2_NAME)
        ws2.views.sheetView[0].showGridLines = True
        self.ws2 = ws2

        headers2 = [
            "Time", "Symbol", "Signal", "Pattern", "Price", "Score",
            "Pivot", "PDH", "PDL", "R1", "S1", "Rel Vol", "Conditions Met"
        ]
        ws2.append(headers2)

        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill_green
            cell.alignment = align_center

        for col in ws2.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 14)

        # Save initial template to disk
        self.save_to_disk()

        # Start live background updating & COM integration
        if self.auto_open:
            self._start_live_updater_thread()

    def update_price(self, symbol: str, ltp: float, volume: int = 0, timestamp: Optional[datetime] = None):
        """
        Updates live price in in-memory openpyxl sheet and queues update for COM thread.
        """
        p = self.pivots.get(symbol)
        pdc = p.pdc if p and p.pdc > 0 else ltp
        change_pct = ((ltp - pdc) / pdc) * 100.0 if pdc > 0 else 0.0
        time_str = (timestamp or datetime.now()).strftime("%H:%M:%S")

        with self._lock:
            # 1. Update in-memory openpyxl sheet
            row_num = self.symbol_row_map.get(symbol)
            if self.ws1 and row_num:
                self.ws1.cell(row=row_num, column=2, value=ltp)
                self.ws1.cell(row=row_num, column=3, value=change_pct / 100.0)
                if volume > 0:
                    self.ws1.cell(row=row_num, column=4, value=volume)
                self.ws1.cell(row=row_num, column=17, value=time_str)

            # 2. Queue for COM updater
            self._live_cache[symbol] = {
                "ltp": ltp,
                "volume": volume,
                "change_pct": change_pct,
                "time": time_str,
            }

    def add_signal(self, signal: SignalEvent):
        """
        Appends an actionable signal row into both in-memory openpyxl Sheet 2 and COM queue.
        """
        ts_str = signal.timestamp.strftime("%H:%M:%S") if isinstance(signal.timestamp, datetime) else str(signal.timestamp)

        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        font_green = Font(name="Calibri", size=10, bold=True, color="006100")
        font_red = Font(name="Calibri", size=10, bold=True, color="9C0006")

        with self._lock:
            # 1. Append directly to in-memory openpyxl Sheet 2
            if self.ws2:
                self.ws2.append([
                    ts_str,
                    signal.symbol,
                    signal.direction,
                    signal.pattern,
                    round(signal.price, 2),
                    signal.score,
                    round(signal.pivot, 2),
                    round(signal.pdh, 2),
                    round(signal.pdl, 2),
                    round(signal.r1, 2),
                    round(signal.s1, 2),
                    round(signal.relative_volume, 2),
                    ", ".join(signal.conditions_met),
                ])
                row_idx = self.ws2.max_row
                cell_sig = self.ws2.cell(row=row_idx, column=3)
                if "BULLISH" in signal.direction:
                    cell_sig.fill = fill_green
                    cell_sig.font = font_green
                elif "BEARISH" in signal.direction:
                    cell_sig.fill = fill_red
                    cell_sig.font = font_red

            # 2. Queue for COM
            self._signals_cache.append(signal)

    def save_to_disk(self):
        """
        Saves the current in-memory openpyxl workbook directly to disk.
        """
        if not self.wb:
            return
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            self.wb.save(self.file_path)
            logger.debug(f"Excel workbook saved to: {self.file_path.name}")
        except PermissionError:
            # If Excel currently has the file open and locked, save to parallel updated file
            alt_path = self.file_path.with_name(f"{self.file_path.stem}_all_signals.xlsx")
            try:
                self.wb.save(alt_path)
                logger.info(f"Primary file open in Excel. Saved updated copy to: {alt_path.name}")
            except Exception:
                pass
        except Exception as e:
            logger.debug(f"Could not save workbook to disk: {e}")

    def _start_live_updater_thread(self):
        """
        Launches background thread to push real-time updates directly into open Excel via COM automation.
        """
        self._stop_com_thread = False
        self._bg_thread = threading.Thread(target=self._com_updater_loop, daemon=True)
        self._bg_thread.start()

    def _com_updater_loop(self):
        """
        Background loop using win32com to update open Excel cells live.
        """
        import pythoncom
        pythoncom.CoInitialize()

        try:
            import win32com.client
            try:
                self._excel_app = win32com.client.Dispatch("Excel.Application")
            except Exception as e:
                logger.debug(f"Could not dispatch Excel application: {e}")
                return

            try:
                self._excel_app.Visible = True
            except Exception:
                pass

            try:
                self._excel_app.DisplayAlerts = False
            except Exception:
                pass

            # Check if workbook is already open in Excel, else Open it
            abs_path = str(self.file_path.resolve())
            wb_found = None
            try:
                for open_wb in self._excel_app.Workbooks:
                    if open_wb.FullName.lower() == abs_path.lower() or open_wb.Name.lower() == self.file_path.name.lower():
                        wb_found = open_wb
                        break
            except Exception:
                pass

            try:
                if wb_found:
                    self._workbook_com = wb_found
                else:
                    self._workbook_com = self._excel_app.Workbooks.Open(abs_path)

                self._sheet1_com = self._workbook_com.Worksheets(self.SHEET1_NAME)
                self._sheet2_com = self._workbook_com.Worksheets(self.SHEET2_NAME)
                self._com_initialized = True
                logger.info("Live Excel window connected via COM automation.")
            except Exception as e:
                logger.debug(f"Could not attach sheets via COM: {e}")
                return

            next_signal_row = 2
            try:
                next_signal_row = self._sheet2_com.UsedRange.Rows.Count + 1
                if next_signal_row < 2:
                    next_signal_row = 2
            except Exception:
                next_signal_row = 2

            while not self._stop_com_thread:
                # 1. Flush Price Updates to COM
                updates_to_push = {}
                with self._lock:
                    if self._live_cache:
                        updates_to_push = dict(self._live_cache)
                        self._live_cache.clear()

                if updates_to_push:
                    for sym, data in updates_to_push.items():
                        row_num = self.symbol_row_map.get(sym)
                        if row_num:
                            try:
                                self._sheet1_com.Cells(row_num, 2).Value = data["ltp"]
                                self._sheet1_com.Cells(row_num, 3).Value = data["change_pct"] / 100.0
                                if data["volume"] > 0:
                                    self._sheet1_com.Cells(row_num, 4).Value = data["volume"]
                                self._sheet1_com.Cells(row_num, 17).Value = data["time"]
                            except Exception:
                                pass

                # 2. Flush Signals to COM
                signals_to_push = []
                with self._lock:
                    if self._signals_cache:
                        signals_to_push = list(self._signals_cache)
                        self._signals_cache.clear()

                if signals_to_push:
                    for sig in signals_to_push:
                        try:
                            ts_str = sig.timestamp.strftime("%H:%M:%S") if isinstance(sig.timestamp, datetime) else str(sig.timestamp)
                            self._sheet2_com.Cells(next_signal_row, 1).Value = ts_str
                            self._sheet2_com.Cells(next_signal_row, 2).Value = sig.symbol
                            self._sheet2_com.Cells(next_signal_row, 3).Value = sig.direction
                            self._sheet2_com.Cells(next_signal_row, 4).Value = sig.pattern
                            self._sheet2_com.Cells(next_signal_row, 5).Value = sig.price
                            self._sheet2_com.Cells(next_signal_row, 6).Value = sig.score
                            self._sheet2_com.Cells(next_signal_row, 7).Value = sig.pivot
                            self._sheet2_com.Cells(next_signal_row, 8).Value = sig.pdh
                            self._sheet2_com.Cells(next_signal_row, 9).Value = sig.pdl
                            self._sheet2_com.Cells(next_signal_row, 10).Value = sig.r1
                            self._sheet2_com.Cells(next_signal_row, 11).Value = sig.s1
                            self._sheet2_com.Cells(next_signal_row, 12).Value = sig.relative_volume
                            self._sheet2_com.Cells(next_signal_row, 13).Value = ", ".join(sig.conditions_met)

                            # Soft color formatting on Signal column
                            cell_sig = self._sheet2_com.Cells(next_signal_row, 3)
                            if "BULLISH" in sig.direction:
                                cell_sig.Interior.Color = 0xC6EFCE  # Soft Light Green
                                cell_sig.Font.Color = 0x006100      # Dark Green Text
                            elif "BEARISH" in sig.direction:
                                cell_sig.Interior.Color = 0xFFC7CE  # Soft Light Red
                                cell_sig.Font.Color = 0x9C0006      # Dark Red Text

                            next_signal_row += 1
                        except Exception:
                            pass

                time.sleep(config.EXCEL_UPDATE_INTERVAL_SECONDS)

        except Exception as e:
            logger.debug(f"Excel COM background loop note: {e}")
        finally:
            pythoncom.CoUninitialize()

    def close(self):
        """Stops live updater and saves in-memory workbook to disk."""
        self._stop_com_thread = True
        if self._bg_thread and self._bg_thread.is_alive():
            self._bg_thread.join(timeout=2.0)

        # Final disk save
        self.save_to_disk()
