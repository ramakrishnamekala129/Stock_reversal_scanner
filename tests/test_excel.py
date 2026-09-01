"""
Unit tests for Live Excel Export Manager.
"""

from datetime import datetime
from pathlib import Path
import openpyxl
from excel.live_excel import LiveExcelManager
from indicators.pivots import calculate_daily_pivots
from scanner.signal_engine import SignalEvent


def test_excel_workbook_initialization(tmp_path: Path):
    excel_file = tmp_path / "test_scanner.xlsx"
    mgr = LiveExcelManager(file_path=excel_file, auto_open=False)

    pivots = {
        "RELIANCE": calculate_daily_pivots("RELIANCE", "2026-08-31", 1280.0, 1300.0, 1270.0, 1290.0, 1000000),
        "TCS": calculate_daily_pivots("TCS", "2026-08-31", 3500.0, 3550.0, 3450.0, 3500.0, 500000),
    }

    mgr.initialize_workbook(pivots)

    assert excel_file.exists()

    wb = openpyxl.load_workbook(excel_file)
    assert LiveExcelManager.SHEET1_NAME in wb.sheetnames
    assert LiveExcelManager.SHEET2_NAME in wb.sheetnames

    ws1 = wb[LiveExcelManager.SHEET1_NAME]
    # Check headers
    assert ws1.cell(row=1, column=1).value == "Symbol"
    assert ws1.cell(row=1, column=2).value == "LTP"
    assert ws1.cell(row=1, column=10).value == "Pivot (PP)"

    # Check rows populated
    symbols_in_sheet = [ws1.cell(row=r, column=1).value for r in range(2, ws1.max_row + 1)]
    assert "RELIANCE" in symbols_in_sheet
    assert "TCS" in symbols_in_sheet
